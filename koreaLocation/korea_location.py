# -*- coding: utf-8 -*- 
import os
import database as db
from openpyxl import load_workbook


def excelTest(MySQLdb):
    path_dir = 'C:\\koreaLocation\\XXYY_DATA.xlsx'   # 가져올 파일들이 있는 directory path
   #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook(path_dir, data_only=True)
    #시트 이름으로 불러오기
    load_ws = load_wb['sheet1']
     
    all_values = []
    for row in load_ws.rows:
        row_value = []
        idx = 0
        for cell in row:
            idx += 1
            if( idx != 6):
                row_value.append(cell.value)
        all_values.append(row_value)
    
    #print(all_values)
    for value in all_values:
        print(value)
        sql = "REPLACE INTO KOREA_LOCATION(LOCATION_A, LOCATION_B, LOCATION_C, X, Y ) VALUES (%s, %s, %s, %s, %s)"
        params = [( value[0], 
                    value[1],
                    value[2],
                    value[3],
                    value[4]
        )] 
        MySQLdb.insert_CQMS_TABLE(sql, params, 'location')     
    #셀 좌표로 값 출력
    #print(load_ws.cell(1,2).value)
        
    

if __name__== "__main__":
    MySQLdb = db.database()
    MySQLdb.version_Select()
    excelTest(MySQLdb);